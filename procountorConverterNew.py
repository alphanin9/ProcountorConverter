from dateutil.parser import parse as parse_date
import dearpygui.dearpygui as imgui

import openpyxl
import csv
import os

from pathlib import Path

imgui.create_context()

def error_window(err_desc: str):
    with imgui.window(label="Virhe"):
        imgui.add_text(err_desc)


def is_revolut(filename):
    return os.path.splitext(filename)[1] == ".csv"


def kill(exitcode):
    imgui.destroy_context()
    exit(exitcode)


def load_revolut(filepath):
    creditAcct = imgui.get_value("__revolut_credit_account")
    topupAcct = imgui.get_value("__revolut_topup_account")
    debitAcct = imgui.get_value("__revolut_debit_account")

    data = []

    with open(filepath) as file:
        reader = csv.reader(file)
        next(reader)
        for line in reader:
            if len(line) < 3:
                continue
            dataEntry = {}

            isTopup = line[3] == "TOPUP"

            dataEntry["name"] = line[5]
            dataEntry["amount"] = float(line[12])
            dataEntry["credit"] = debitAcct if isTopup else creditAcct
            dataEntry["debit"] = topupAcct if isTopup else debitAcct
            dataEntry["date"] = parse_date(line[0], yearfirst=True)

            data.append(dataEntry)
        pass

    return data


def read_workbook(filepath):
    workbook = openpyxl.load_workbook(filename=filepath)

    if not workbook:
        kill(1)

    worksheet = workbook["in"]
    row_data = []
    for row in worksheet.iter_rows(values_only=True, max_col=6):
        print(row)
    
    for row in worksheet.iter_rows(values_only=True, max_col=6):
        used_date = row[0]
        
        if isinstance(row[0], str):
            try:
                used_date = parse_date(row[0], dayfirst=True)
            except:
                print(f"Something went wrong with row {row} date parsing! Skipping row...")
                continue
        
        row_data.append(
            {
                "name": row[1],
                "credit": row[2],
                "debit": row[3],
                "amount": row[5] or row[4],
                "date": used_date,
            }
        )

    return row_data


def write_row_data(filepath, row_data):
    by_date = {}
    for i in row_data:
        if i["date"] not in by_date:
            by_date[i["date"]] = []
        by_date[i["date"]].append(i)

    root_path = Path(os.path.splitext(filepath)[0] + "_OUTPUT")
    
    if not root_path.is_dir():
        root_path.mkdir()
    
    for i in by_date.keys():
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        worksheet.title = "out"

        for row_raw_data in by_date[i]:
            price = row_raw_data["amount"]

            row_1 = (row_raw_data["credit"], row_raw_data["name"], price)
            row_2 = (row_raw_data["debit"], row_raw_data["name"], price * -1)

            worksheet.append(row_1)
            worksheet.append(row_2)

        new_filename = root_path / f"{i.strftime('%d-%m-%Y')}.xlsx"
        workbook.save(new_filename)


def main(filepath):
    row_data = []

    if is_revolut(filepath):
        row_data = load_revolut(filepath)
    else:
        row_data = read_workbook(filepath)

    if not row_data:
        kill(1)
        return

    write_row_data(filepath, row_data)


def file_selection_callback(sender, appdata):
    main(appdata["file_path_name"])
    kill(0)


with imgui.file_dialog(
    directory_selector=False,
    show=False,
    tag="open_file_dialog",
    width=700,
    height=500,
    callback=file_selection_callback,
):
    imgui.add_file_extension(".xlsx")
    imgui.add_file_extension(".csv")

with imgui.window(label="Converter", width=800, height=600, no_title_bar=True):
    imgui.add_button(
        label="Valitse tiedosto...",
        callback=lambda: imgui.show_item("open_file_dialog"),
    )
    imgui.add_input_int(
        label="Revolut credit account",
        tag="__revolut_credit_account",
        default_value=2880,
    )
    imgui.add_input_int(
        label="Revolut topup account", tag="__revolut_topup_account", default_value=1700
    )
    imgui.add_input_int(
        label="Revolut debit account", tag="__revolut_debit_account", default_value=1930
    )

imgui.create_viewport(title="Converter", width=800, height=600)
imgui.setup_dearpygui()
imgui.show_viewport()
imgui.start_dearpygui()
imgui.destroy_context()
